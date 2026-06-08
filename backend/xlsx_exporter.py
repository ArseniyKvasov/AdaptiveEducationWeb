from __future__ import annotations

import re
import logging
from io import BytesIO
from datetime import datetime
from typing import Any, Optional

from .pipeline import speech_analysis_from_generation, format_timestamp, transcript_lines_by_ms_range

logger = logging.getLogger(__name__)


def sanitize_xlsx_sheet_name(name: str, fallback: str = "Лист") -> str:
    """Removes invalid characters and bounds length to 31 chars for spreadsheet tabs."""
    cleaned = re.sub(r"[\[\]\*\/\\\?:]", " ", str(name or fallback))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = cleaned[:31]
    return cleaned or fallback


def normalize_speech_title(title: Any) -> str:
    """Provides shorter localized translations of specific analysis titles."""
    value = str(title or "").strip()
    if value == "Преподаватель активно задаёт вопросы студентам":
        return "Вопросы преподавателя"
    if value == "Преподаватель реагирует на ответы студентов и развивает обсуждение":
        return "Ответы студентов"
    return value


def normalize_speech_analysis_export_view(speech: Any) -> dict[str, Any]:
    """Prepares and merges fragment results across audio chunks into consolidated worksheets."""
    question_type_aliases = {
        "rhetorical": "rhetorical",
        "риторический": "rhetorical",
        "checking_understanding": "checking_understanding",
        "проверка понимания": "checking_understanding",
        "quiz": "quiz",
        "викторина": "quiz",
        "clarifying": "clarifying",
        "уточняющий": "clarifying",
        "open_ended": "open_ended",
        "open-ended": "open_ended",
        "открытый": "open_ended",
        "factual": "factual",
        "фактический": "factual",
        "other": "other",
        "другой": "other",
    }

    def normalize_question_type(value: Any) -> str:
        normalized = str(value or "").strip().casefold()
        return question_type_aliases.get(normalized, "other" if normalized else "")

    def normalize_fragment(fragment: Any) -> dict[str, Any]:
        if not isinstance(fragment, dict):
            text = str(fragment or "").strip()
            return {"start_ms": 0, "end_ms": 0, "text": text}
        try:
            start_value = int(fragment.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(fragment.get("end_ms", start_value) or start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            end_value = start_value
        normalized: dict[str, Any] = {
            "start_ms": start_value,
            "end_ms": end_value,
            "text": str(fragment.get("text") or "").strip(),
        }
        fragment_type = str(fragment.get("type") or "").strip()
        if fragment_type:
            normalized["type"] = fragment_type
        question_type = normalize_question_type(fragment.get("question_type"))
        if question_type:
            normalized["question_type"] = question_type
        return normalized

    def merge_fragments(primary: list[dict[str, Any]], secondary: list[dict[str, Any]]) -> list[dict[str, Any]]:
        merged: list[dict[str, Any]] = []
        seen: set[tuple[int, int, str, str, str]] = set()
        for fragment in primary + secondary:
            if not isinstance(fragment, dict):
                continue
            key = (
                int(fragment.get("start_ms", 0) or 0),
                int(fragment.get("end_ms", 0) or 0),
                str(fragment.get("text") or "").strip(),
                str(fragment.get("type") or "").strip(),
                normalize_question_type(fragment.get("question_type")),
            )
            if not key[2] or key in seen:
                continue
            seen.add(key)
            normalized = dict(fragment)
            qtype = normalize_question_type(normalized.get("question_type"))
            if qtype:
                normalized["question_type"] = qtype
            elif "question_type" in normalized:
                normalized.pop("question_type", None)
            merged.append(normalized)
        return merged

    def collect_chunk_fragments(chunk_analyses: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
        fragments: list[dict[str, Any]] = []
        for chunk in chunk_analyses:
            if not isinstance(chunk, dict):
                continue
            items = chunk.get(key)
            if not isinstance(items, list):
                continue
            fragments.extend(normalize_fragment(item) for item in items if isinstance(item, dict) and str(item.get("text") or "").strip())
        return fragments

    def collect_chunk_events(chunk_analyses: list[dict[str, Any]]) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for chunk in chunk_analyses:
            if not isinstance(chunk, dict):
                continue
            events = chunk.get("lesson_events")
            if not isinstance(events, list):
                continue
            for event in events:
                if not isinstance(event, dict):
                    continue
                try:
                    start_value = int(event.get("start_ms", 0) or 0)
                except (TypeError, ValueError):
                    start_value = 0
                items.append(
                    {
                        "time_ms": start_value,
                        "time": format_timestamp(start_value / 1000),
                        "title": str(event.get("title") or "").strip() or "Событие урока",
                        "description": str(event.get("description") or "").strip(),
                    }
                )
        items.sort(key=lambda item: int(item.get("time_ms", 0) or 0))
        return items

    speech_data = speech if isinstance(speech, dict) else {}
    chunk_analyses = speech_data.get("chunk_analyses") if isinstance(speech_data.get("chunk_analyses"), list) else []

    lesson_format_raw = speech_data.get("lesson_format") if isinstance(speech_data.get("lesson_format"), dict) else {}
    audience_engagement_raw = speech_data.get("audience_engagement") if isinstance(speech_data.get("audience_engagement"), dict) else {}
    lesson_structure_raw = speech_data.get("lesson_structure") if isinstance(speech_data.get("lesson_structure"), dict) else {}
    material_explanation_raw = speech_data.get("material_explanation") if isinstance(speech_data.get("material_explanation"), dict) else {}
    teacher_recommendation_raw = speech_data.get("teacher_recommendation") if isinstance(speech_data.get("teacher_recommendation"), dict) else {}
    flags_raw = speech_data.get("flags") if isinstance(speech_data.get("flags"), dict) else {}

    questions_raw = audience_engagement_raw.get("questions_to_students") if isinstance(audience_engagement_raw.get("questions_to_students"), dict) else {}
    answers_raw = audience_engagement_raw.get("student_answers") if isinstance(audience_engagement_raw.get("student_answers"), dict) else {}
    timeline_raw = lesson_structure_raw.get("step_by_step_explanation") if isinstance(lesson_structure_raw.get("step_by_step_explanation"), dict) else {}
    goals_raw = lesson_structure_raw.get("goals_and_summary") if isinstance(lesson_structure_raw.get("goals_and_summary"), dict) else {}
    examples_raw = material_explanation_raw.get("examples_and_analogies") if isinstance(material_explanation_raw.get("examples_and_analogies"), dict) else {}

    derived_questions = collect_chunk_fragments(chunk_analyses, "teacher_questions")
    derived_answers = collect_chunk_fragments(chunk_analyses, "student_answers")
    derived_examples = collect_chunk_fragments(chunk_analyses, "examples_and_analogies")
    derived_timeline = collect_chunk_events(chunk_analyses)
    derived_profanity: list[dict[str, Any]] = []
    derived_familiarity: list[dict[str, Any]] = []
    for chunk in chunk_analyses:
        if not isinstance(chunk, dict):
            continue
        chunk_flags = chunk.get("flags") if isinstance(chunk.get("flags"), dict) else {}
        profanity_items = chunk_flags.get("profanity") if isinstance(chunk_flags.get("profanity"), list) else []
        familiarity_items = chunk_flags.get("overly_familiar_tone") if isinstance(chunk_flags.get("overly_familiar_tone"), list) else []
        derived_profanity.extend(normalize_fragment(item) for item in profanity_items if isinstance(item, dict) and str(item.get("text") or "").strip())
        derived_familiarity.extend(normalize_fragment(item) for item in familiarity_items if isinstance(item, dict) and str(item.get("text") or "").strip())

    derived_intro: dict[str, Any] | None = None
    derived_summary: dict[str, Any] | None = None
    for chunk in chunk_analyses:
        if not isinstance(chunk, dict):
            continue
        chunk_goals = chunk.get("goals_and_summary") if isinstance(chunk.get("goals_and_summary"), dict) else {}
        intro = chunk_goals.get("intro") if isinstance(chunk_goals.get("intro"), dict) else {}
        summary = chunk_goals.get("summary") if isinstance(chunk_goals.get("summary"), dict) else {}
        if derived_intro is None and intro:
            derived_intro = {
                "present": bool(intro.get("present")),
                "start_ms": intro.get("start_ms"),
                "comment": str(intro.get("comment") or "").strip(),
            }
        elif derived_intro is not None and not derived_intro.get("present") and bool(intro.get("present")):
            derived_intro = {
                "present": True,
                "start_ms": intro.get("start_ms"),
                "comment": str(intro.get("comment") or "").strip(),
            }
        if derived_summary is None and summary:
            derived_summary = {
                "present": bool(summary.get("present")),
                "start_ms": summary.get("start_ms"),
                "comment": str(summary.get("comment") or "").strip(),
            }
        elif derived_summary is not None and not derived_summary.get("present") and bool(summary.get("present")):
            derived_summary = {
                "present": True,
                "start_ms": summary.get("start_ms"),
                "comment": str(summary.get("comment") or "").strip(),
            }

    lesson_format = {
        "format": str(lesson_format_raw.get("format") or "").strip() or (
            "Агрегированный анализ речи преподавателя"
            if chunk_analyses
            else "Формат занятия не определен"
        ),
        "comment": str(lesson_format_raw.get("comment") or "").strip() or (
            f"Проанализировано чанков: {len(chunk_analyses)}"
            if chunk_analyses
            else "Агрегированный анализ речи преподавателя готов."
        ),
    }

    questions = {
        "title": normalize_speech_title(questions_raw.get("title") or "Вопросы преподавателя"),
        "comment": str(questions_raw.get("comment") or "").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (questions_raw.get("fragments") if isinstance(questions_raw.get("fragments"), list) else [])],
            derived_questions,
        ),
    }
    answers = {
        "title": normalize_speech_title(answers_raw.get("title") or "Ответы студентов"),
        "comment": str(answers_raw.get("comment") or "").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (answers_raw.get("fragments") if isinstance(answers_raw.get("fragments"), list) else [])],
            derived_answers,
        ),
    }

    timeline = {
        "title": str(timeline_raw.get("title") or "Таймлайн урока").strip(),
        "timeline": [],
    }
    if isinstance(timeline_raw.get("timeline"), list) and timeline_raw.get("timeline"):
        for item in timeline_raw.get("timeline"):
            if not isinstance(item, dict):
                continue
            try:
                start_value = int(item.get("start_ms", 0) or 0)
            except (TypeError, ValueError):
                start_value = 0
            timeline["timeline"].append(
                {
                    "start_ms": start_value,
                    "time": format_timestamp(start_value / 1000),
                    "title": str(item.get("title") or "Событие урока").strip(),
                    "comment": str(item.get("description") or "").strip(),
                }
            )
    else:
        timeline["timeline"] = [
            {"start_ms": item["time_ms"], "time": item["time"], "title": item["title"], "comment": item["description"]}
            for item in derived_timeline
        ]

    intro_raw = goals_raw.get("intro") if isinstance(goals_raw.get("intro"), dict) else {}
    summary_raw = goals_raw.get("summary") if isinstance(goals_raw.get("summary"), dict) else {}
    goals = {
        "title": str(goals_raw.get("title") or "Цели и итоги урока").strip(),
        "intro": {
            "present": bool(intro_raw.get("present")) if intro_raw else bool(derived_intro.get("present")) if isinstance(derived_intro, dict) else False,
            "start_ms": intro_raw.get("start_ms") if intro_raw else (derived_intro.get("start_ms") if isinstance(derived_intro, dict) else None),
            "comment": str(intro_raw.get("comment") or "").strip() if intro_raw else (str(derived_intro.get("comment") or "").strip() if isinstance(derived_intro, dict) else ""),
        },
        "summary": {
            "present": bool(summary_raw.get("present")) if summary_raw else bool(derived_summary.get("present")) if isinstance(derived_summary, dict) else False,
            "start_ms": summary_raw.get("start_ms") if summary_raw else (derived_summary.get("start_ms") if isinstance(derived_summary, dict) else None),
            "comment": str(summary_raw.get("comment") or "").strip() if summary_raw else (str(derived_summary.get("comment") or "").strip() if isinstance(derived_summary, dict) else ""),
        },
    }

    examples = {
        "title": str(examples_raw.get("title") or "Примеры, аналогии и сторителлинг").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (examples_raw.get("fragments") if isinstance(examples_raw.get("fragments"), list) else [])],
            derived_examples,
        ),
    }

    recommendation = {
        "title": str(teacher_recommendation_raw.get("title") or "Рекомендация преподавателю").strip(),
        "comment": str(teacher_recommendation_raw.get("comment") or "").strip(),
    }

    flags: dict[str, Any] = {}
    for key, fallback_title, derived_fragments in [
        ("profanity", "Ненормативная лексика", derived_profanity),
        ("overly_familiar_tone", "Панибратство", derived_familiarity),
    ]:
        block = flags_raw.get(key) if isinstance(flags_raw.get(key), dict) else {}
        fragments = block.get("fragments") if isinstance(block.get("fragments"), list) else []
        flags[key] = {
            "title": str(block.get("title") or fallback_title).strip(),
            "present": bool(block.get("present")) or bool(derived_fragments) or bool(fragments),
            "fragments": merge_fragments(
                [normalize_fragment(fragment) for fragment in fragments],
                derived_fragments,
            ),
        }

    return {
        "lesson_format": lesson_format,
        "audience_engagement": {
            "questions_to_students": questions,
            "student_answers": answers,
        },
        "lesson_structure": {
            "step_by_step_explanation": timeline,
            "goals_and_summary": goals,
        },
        "material_explanation": {
            "examples_and_analogies": examples,
        },
        "teacher_recommendation": recommendation,
        "flags": flags,
        "chunk_analyses": chunk_analyses,
    }


def build_speech_analysis_export_worksheets_precise(generation: dict[str, Any]) -> list[dict[str, Any]]:
    """Builds worksheets formatted for teacher report download, with precise ratings and comments."""
    transcript = generation.get("transcript") if isinstance(generation.get("transcript"), list) else []
    speech = speech_analysis_from_generation(generation)
    if not speech:
        return []
    speech = normalize_speech_analysis_export_view(speech)

    lesson_format = speech.get("lesson_format") if isinstance(speech.get("lesson_format"), dict) else {}
    audience_engagement = speech.get("audience_engagement") if isinstance(speech.get("audience_engagement"), dict) else {}
    lesson_structure = speech.get("lesson_structure") if isinstance(speech.get("lesson_structure"), dict) else {}
    material_explanation = speech.get("material_explanation") if isinstance(speech.get("material_explanation"), dict) else {}
    flags = speech.get("flags") if isinstance(speech.get("flags"), dict) else {}
    teacher_recommendation = speech.get("teacher_recommendation") if isinstance(speech.get("teacher_recommendation"), dict) else {}

    questions = audience_engagement.get("questions_to_students") if isinstance(audience_engagement.get("questions_to_students"), dict) else {}
    student_answers = audience_engagement.get("student_answers") if isinstance(audience_engagement.get("student_answers"), dict) else {}
    timeline = lesson_structure.get("step_by_step_explanation") if isinstance(lesson_structure.get("step_by_step_explanation"), dict) else {}
    goals = lesson_structure.get("goals_and_summary") if isinstance(lesson_structure.get("goals_and_summary"), dict) else {}
    examples = material_explanation.get("examples_and_analogies") if isinstance(material_explanation.get("examples_and_analogies"), dict) else {}

    questions_fragments = questions.get("fragments") if isinstance(questions.get("fragments"), list) else []
    answers_fragments = student_answers.get("fragments") if isinstance(student_answers.get("fragments"), list) else []
    examples_fragments = examples.get("fragments") if isinstance(examples.get("fragments"), list) else []
    timeline_items = timeline.get("timeline") if isinstance(timeline.get("timeline"), list) else []
    profanity_block = flags.get("profanity") if isinstance(flags.get("profanity"), dict) else {}
    familiarity_block = flags.get("overly_familiar_tone") if isinstance(flags.get("overly_familiar_tone"), dict) else {}
    profanity_fragments = profanity_block.get("fragments") if isinstance(profanity_block.get("fragments"), list) else []
    familiarity_fragments = familiarity_block.get("fragments") if isinstance(familiarity_block.get("fragments"), list) else []

    question_type_labels = {
        "rhetorical": "риторический",
        "checking_understanding": "проверка понимания",
        "quiz": "викторина",
        "clarifying": "уточняющий",
        "open_ended": "открытый",
        "factual": "фактический",
        "other": "другой",
    }
    question_type_order = ["checking_understanding", "open_ended", "clarifying", "quiz", "factual", "rhetorical", "other"]

    def clean(value: Any) -> str:
        return str(value or "").strip()

    def percentage(part: int, total: int) -> str:
        if total <= 0:
            return "0%"
        return f"{round((part / total) * 100)}%"

    def timestamp_range(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        try:
            start_value = int(fragment.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(fragment.get("end_ms", start_value) or start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            start_value, end_value = end_value, start_value
        return f"{format_timestamp(start_value / 1000)}–{format_timestamp(end_value / 1000)}" if end_value > start_value else format_timestamp(start_value / 1000)

    def transcript_text(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        lines = transcript_lines_by_ms_range(transcript, fragment.get("start_ms"), fragment.get("end_ms"))
        def line_text(line: dict[str, Any]) -> str:
            text = clean(line.get("text"))
            if not text:
                return ""
            speaker = clean(line.get("speaker"))
            return f"{speaker}: {text}" if speaker else text
        return " / ".join(line_text(line) for line in lines if line_text(line))

    def fragment_manifestation(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return clean(fragment)
        text = clean(fragment.get("text"))
        ftype = clean(fragment.get("type")).casefold()
        if ftype in {"example", "analogy", "metaphor", "storytelling"}:
            return f"{text} ({ {'example': 'пример', 'analogy': 'аналогия', 'metaphor': 'метафора', 'storytelling': 'сторителлинг'}[ftype] })" if text else ""
        qtype = clean(fragment.get("question_type")).casefold()
        if qtype in question_type_labels and text:
            return f"{text} [{question_type_labels[qtype]}]"
        return text

    def example_type_label(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return "Пример"
        ftype = clean(fragment.get("type")).casefold()
        labels = {
            "example": "Пример",
            "analogy": "Аналогия",
            "metaphor": "Метафора",
            "storytelling": "Сторителлинг",
        }
        return labels.get(ftype, "Пример")

    def fragment_comment(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        parts = [piece for piece in [timestamp_range(fragment), transcript_text(fragment)] if piece]
        return " | ".join(parts)

    def join_comments(fragments: list[dict[str, Any]], limit: int = 3) -> str:
        parts: list[str] = []
        seen: set[str] = set()
        for fragment in fragments[:limit]:
            comment = fragment_comment(fragment)
            if comment and comment not in seen:
                seen.add(comment)
                parts.append(comment)
        return "\n".join(parts)

    def point(_value: int) -> str:
        return "-"

    def question_points(count: int) -> int:
        return 2 if count > 0 else 0

    def atmosphere_points(answers_count: int, questions_count: int) -> int:
        if questions_count <= 0:
            return 0
        ratio = answers_count / questions_count
        if ratio > 0.4:
            return 2
        if ratio > 0.2:
            return 1
        return 0

    created_at = generation.get("created_at")
    date_text = ""
    if created_at:
        try:
            date_text = datetime.fromisoformat(str(created_at).replace("Z", "+00:00")).strftime("%d.%m.%Y")
        except Exception:
            date_text = clean(created_at)

    total_questions = len(questions_fragments)
    total_answers = len(answers_fragments)
    total_examples = len(examples_fragments)
    total_events = len(timeline_items)
    has_intro = bool(goals.get("intro", {}).get("present")) if isinstance(goals.get("intro"), dict) else False
    has_summary = bool(goals.get("summary", {}).get("present")) if isinstance(goals.get("summary"), dict) else False
    has_profanity = bool(profanity_block.get("present"))
    has_familiarity = bool(familiarity_block.get("present"))

    question_type_counts: dict[str, int] = {}
    question_type_fragments: dict[str, list[dict[str, Any]]] = {}
    seen_question_keys: set[str] = set()
    for fragment in questions_fragments:
        if not isinstance(fragment, dict):
            continue
        unique_key = f"{fragment.get('start_ms', '')}|{fragment.get('end_ms', '')}|{clean(fragment.get('text'))}"
        if unique_key in seen_question_keys:
            continue
        seen_question_keys.add(unique_key)
        qtype = clean(fragment.get("question_type")).casefold() or "other"
        question_type_counts[qtype] = question_type_counts.get(qtype, 0) + 1
        question_type_fragments.setdefault(qtype, []).append(fragment)

    rows: list[dict[str, Any]] = [
        {"cells": [{"value": "Чек-лист качества преподавания на занятии", "style": 2, "span": 5}], "height": 24},
        {"cells": [{"value": "Заполняйте только по фактам из анализа. Пустые ячейки лучше оставлять пустыми, если в данных нет подтверждения.", "style": 7, "span": 5}], "height": 34},
        {"cells": [{"value": "Дата формирования", "style": 14, "span": 1}, {"value": date_text, "style": 5, "span": 2}, {"value": "Файл", "style": 14, "span": 1}, {"value": clean(generation.get("file_name")), "style": 5, "span": 1}], "height": 20},
        {"cells": [{"value": "Формат занятия", "style": 14, "span": 1}, {"value": clean(lesson_format.get("format")), "style": 5, "span": 4}], "height": 20},
        {"cells": [], "height": 6},
        {"cells": [{"value": "Легенда", "style": 4, "span": 5}], "height": 20},
        {"cells": [{"value": "Баллы", "style": 1}, {"value": "Смысл", "style": 1}, {"value": "Проявление", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}], "height": 20},
        {"cells": [{"value": "0", "style": 11}, {"value": "Отсутствие или несоответствие", "style": 5}, {"value": "Показатель не подтвержден", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [{"value": "1", "style": 12}, {"value": "Частичное соответствие", "style": 5}, {"value": "Показатель подтвержден частично", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [{"value": "2", "style": 13}, {"value": "Полное соответствие", "style": 5}, {"value": "Показатель подтвержден", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [], "height": 8},
        {"cells": [{"value": "Анализ по индикаторам", "style": 4, "span": 5}], "height": 20},
    ]

    def add_section(title: str, rows_to_add: list[dict[str, Any]]) -> None:
        if not rows_to_add:
            return
        rows.append({"cells": [{"value": title, "style": 8, "span": 5}], "height": 22})
        rows.append({"cells": [{"value": "Компетенция", "style": 1}, {"value": "Поведенческие индикаторы", "style": 1}, {"value": "Проявление", "style": 1}, {"value": "Баллы", "style": 1}, {"value": "Дополнительные комментарии", "style": 1}], "height": 20})
        rows.extend(rows_to_add)

    engagement_rows: list[dict[str, Any]] = []
    for qtype in question_type_order:
        count = question_type_counts.get(qtype, 0)
        if not count:
            continue
        fragments_for_type = question_type_fragments.get(qtype, [])
        engagement_rows.append({
            "cells": [
                {"value": "Вопросы преподавателя", "style": 14},
                {"value": question_type_labels.get(qtype, qtype), "style": 5},
                {"value": f"{count} из {total_questions} ({percentage(count, total_questions)})", "style": 5},
                {"value": point(question_points(count)), "style": 13 if count > 0 else 11},
                {"value": join_comments(fragments_for_type), "style": 5},
            ],
            "height": 26,
        })
    answers_comment = join_comments(answers_fragments) if answers_fragments else "Ответов студентов не было."
    atmosphere_score = atmosphere_points(total_answers, total_questions)
    engagement_rows.append({
        "cells": [
            {"value": "Ответы студентов", "style": 14},
            {"value": "Ответы студентов", "style": 5},
            {"value": f"{total_answers} из {total_questions} ({percentage(total_answers, total_questions)})", "style": 5},
            {"value": point(atmosphere_score), "style": 12 if atmosphere_score == 1 else 13 if atmosphere_score == 2 else 11},
            {"value": answers_comment, "style": 5},
        ],
        "height": 26,
    })
    add_section("Вовлечение аудитории", engagement_rows)

    structure_rows: list[dict[str, Any]] = []
    if total_events:
        timeline_comment = "\n".join(
            f"{clean(item.get('time') or '')} · {clean(item.get('title'))}{(' — ' + clean(item.get('comment'))) if clean(item.get('comment')) else ''}"
            for item in timeline_items
            if isinstance(item, dict)
        )
        structure_rows.append({
            "cells": [
                {"value": "Таймлайн урока", "style": 14},
                {"value": "Последовательность этапов", "style": 5},
                {"value": f"{total_events} событий", "style": 5},
                {"value": point(2 if total_events >= 2 else 1), "style": 13 if total_events >= 2 else 12},
                {"value": timeline_comment, "style": 5},
            ],
            "height": 34,
        })
    structure_rows.append({
        "cells": [
            {"value": "Цели и итоги урока", "style": 14},
            {"value": "Введение", "style": 5},
            {"value": "есть" if has_intro else "нет", "style": 5},
            {"value": point(2 if has_intro else 0), "style": 13 if has_intro else 11},
            {"value": clean(goals.get("intro", {}).get("comment") if isinstance(goals.get("intro"), dict) else ""), "style": 5},
        ],
        "height": 24,
    })
    structure_rows.append({
        "cells": [
            {"value": "Цели и итоги урока", "style": 14},
            {"value": "Завершение", "style": 5},
            {"value": "есть" if has_summary else "нет", "style": 5},
            {"value": point(2 if has_summary else 0), "style": 13 if has_summary else 11},
            {"value": clean(goals.get("summary", {}).get("comment") if isinstance(goals.get("summary"), dict) else ""), "style": 5},
        ],
        "height": 24,
    })
    add_section("Структура занятия", structure_rows)

    explanation_rows: list[dict[str, Any]] = []
    for fragment in examples_fragments:
        if not isinstance(fragment, dict):
            continue
        explanation_rows.append({
            "cells": [
                {"value": example_type_label(fragment), "style": 14},
                {"value": "Пример из речи преподавателя", "style": 5},
                {"value": fragment_manifestation(fragment), "style": 5},
                {"value": "-", "style": 5},
                {"value": fragment_comment(fragment), "style": 5},
            ],
            "height": 26,
        })
    add_section("Примеры, аналогии и сторителлинг", explanation_rows)

    flag_rows: list[dict[str, Any]] = []
    if has_profanity or profanity_fragments:
        flag_rows.append({
            "cells": [
                {"value": "Ненормативная лексика", "style": 14},
                {"value": "Тон и лексика", "style": 5},
                {"value": "не подтверждено" if not has_profanity else "есть сигналы", "style": 5},
                {"value": point(2 if not has_profanity and not profanity_fragments else 0), "style": 13 if not has_profanity and not profanity_fragments else 11},
                {"value": join_comments(profanity_fragments), "style": 5},
            ],
            "height": 26,
        })
    if has_familiarity or familiarity_fragments:
        flag_rows.append({
            "cells": [
                {"value": "Панибратство", "style": 14},
                {"value": "Обращение к аудитории", "style": 5},
                {"value": "не подтверждено" if not has_familiarity else "есть сигналы", "style": 5},
                {"value": point(2 if not has_familiarity and not familiarity_fragments else 0), "style": 13 if not has_familiarity and not familiarity_fragments else 11},
                {"value": join_comments(familiarity_fragments), "style": 5},
            ],
            "height": 26,
        })
    add_section("Флаги", flag_rows)

    recommendation_comment = clean(teacher_recommendation.get("comment"))
    recommendation_rows: list[dict[str, Any]] = []
    if recommendation_comment:
        recommendation_rows.append({
            "cells": [
                {"value": "Рекомендация преподавателю", "style": 14},
                {"value": "Комментарий", "style": 5},
                {"value": recommendation_comment, "style": 5},
                {"value": "", "style": 5},
                {"value": "", "style": 5},
            ],
            "height": 30,
        })
    add_section("Рекомендация преподавателю", recommendation_rows)

    return [
        {
            "name": "Анализ речи",
            "rows": rows,
            "cols": [22, 28, 34, 12, 44],
            "page_setup": {"orientation": "landscape", "paperSize": 9, "fitToWidth": 1, "fitToHeight": 1},
            "freeze_panes": "A12",
        }
    ]


def build_speech_analysis_export_worksheets(generation: dict[str, Any]) -> list[dict[str, Any]]:
    """Builds additional aggregate analysis worksheets based on speech flags, timeline, and dynamic metrics."""
    transcript = generation.get("transcript") if isinstance(generation.get("transcript"), list) else []
    speech = speech_analysis_from_generation(generation)
    if not speech:
        return []
    speech = normalize_speech_analysis_export_view(speech)

    question_type_labels = {
        "rhetorical": "риторический",
        "checking_understanding": "проверка понимания",
        "quiz": "викторина",
        "clarifying": "уточняющий",
        "open_ended": "открытый",
        "factual": "фактический",
        "other": "другой",
    }

    def clean(value: Any) -> str:
        return str(value or "").strip()

    def transcript_fragment_row(section: str, title: str, fragment: Any) -> list[str]:
        fragment_text = str(fragment.get("text") or "").strip() if isinstance(fragment, dict) else str(fragment or "").strip()
        if isinstance(fragment, dict):
            fragment_type = str(fragment.get("type") or "").strip().casefold()
            if fragment_type in {"example", "analogy", "metaphor", "storytelling"}:
                type_labels = {"example": "пример", "analogy": "аналогия", "metaphor": "метафора", "storytelling": "сторителлинг"}
                fragment_text = f"{fragment_text} ({type_labels[fragment_type]})"
            qtype = str(fragment.get("question_type") or "").strip().casefold()
            if qtype in question_type_labels:
                fragment_text = f"{fragment_text} [{question_type_labels[qtype]}]"
        start_ms_value: Any = fragment.get("start_ms") if isinstance(fragment, dict) else None
        end_ms_value: Any = fragment.get("end_ms") if isinstance(fragment, dict) else None
        lines = transcript_lines_by_ms_range(transcript, start_ms_value, end_ms_value)
        found = "Да" if lines else "Нет"
        try:
            start_value = int(start_ms_value or 0) if start_ms_value is not None else 0
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(end_ms_value if end_ms_value is not None else start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            start_value, end_value = end_value, start_value
        timestamp = f"{format_timestamp(start_value / 1000)}–{format_timestamp(end_value / 1000)}" if end_value > start_value else format_timestamp(start_value / 1000)
        t_text = " / ".join(
            (f"{line.get('speaker')}: {line.get('text')}" if line.get("speaker") else str(line.get("text") or "")).strip()
            for line in lines
            if str(line.get("text") or "").strip()
        )
        return [section, title, fragment_text, found, timestamp, t_text]

    lesson_format = speech.get("lesson_format") if isinstance(speech.get("lesson_format"), dict) else {}
    audience_engagement = speech.get("audience_engagement") if isinstance(speech.get("audience_engagement"), dict) else {}
    lesson_structure = speech.get("lesson_structure") if isinstance(speech.get("lesson_structure"), dict) else {}
    material_explanation = speech.get("material_explanation") if isinstance(speech.get("material_explanation"), dict) else {}
    flags = speech.get("flags") if isinstance(speech.get("flags"), dict) else {}
    chunk_analyses = speech.get("chunk_analyses") if isinstance(speech.get("chunk_analyses"), list) else []

    questions = audience_engagement.get("questions_to_students") if isinstance(audience_engagement.get("questions_to_students"), dict) else {}
    student_answers = audience_engagement.get("student_answers") if isinstance(audience_engagement.get("student_answers"), dict) else {}
    timeline = lesson_structure.get("step_by_step_explanation") if isinstance(lesson_structure.get("step_by_step_explanation"), dict) else {}
    goals = lesson_structure.get("goals_and_summary") if isinstance(lesson_structure.get("goals_and_summary"), dict) else {}
    examples = material_explanation.get("examples_and_analogies") if isinstance(material_explanation.get("examples_and_analogies"), dict) else {}
    teacher_recommendation = speech.get("teacher_recommendation") if isinstance(speech.get("teacher_recommendation"), dict) else {}

    questions_fragments = questions.get("fragments") if isinstance(questions.get("fragments"), list) else []
    answers_fragments = student_answers.get("fragments") if isinstance(student_answers.get("fragments"), list) else []
    examples_fragments = examples.get("fragments") if isinstance(examples.get("fragments"), list) else []
    timeline_items = timeline.get("timeline") if isinstance(timeline.get("timeline"), list) else []
    profanity_block = flags.get("profanity") if isinstance(flags.get("profanity"), dict) else {}
    familiarity_block = flags.get("overly_familiar_tone") if isinstance(flags.get("overly_familiar_tone"), dict) else {}
    profanity_fragments = profanity_block.get("fragments") if isinstance(profanity_block.get("fragments"), list) else []
    familiarity_fragments = familiarity_block.get("fragments") if isinstance(familiarity_block.get("fragments"), list) else []

    def fmt_score(_value: float) -> str:
        return "-"

    def score_style(value: float) -> int:
        if value >= 1.5:
            return 13
        if value >= 0.5:
            return 12
        return 11

    def factual_label(score: float) -> str:
        if score >= 1.5:
            return "было"
        if score >= 0.5:
            return "частично"
        return "не было"

    def count_types(fragments: list[dict[str, Any]], types: set[str]) -> int:
        total = 0
        for fragment in fragments:
            if not isinstance(fragment, dict):
                continue
            if str(fragment.get("question_type") or "").strip().casefold() in types:
                total += 1
        return total

    def section_rows(title: str, fill_style: int, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        block: list[dict[str, Any]] = [{"cells": [{"value": title, "style": fill_style, "span": 6}], "height": 22}]
        block.append({"cells": [{"value": h, "style": 1} for h in ["Компетенция", "Поведенческие индикаторы", "Проявление", "Фактические действия (было/не было)", "Средний балл", "Дополнительные комментарии"]], "height": 20})
        scores: list[float] = []
        for item in items:
            score = float(item.get("score", 0) or 0)
            scores.append(score)
            block.append({
                "cells": [
                    {"value": item.get("competence", ""), "style": 14, "span": 1},
                    {"value": item.get("indicator", ""), "style": 5, "span": 1},
                    {"value": item.get("manifestation", ""), "style": 5, "span": 1},
                    {"value": factual_label(score), "style": 5, "span": 1},
                    {"value": fmt_score(score), "style": score_style(score), "span": 1},
                    {"value": item.get("comment", ""), "style": 5, "span": 1},
                ],
                "height": int(item.get("height", 24) or 24),
            })
        avg = sum(scores) / len(scores) if scores else 0
        note = str(items[-1].get("summary_note") or "") if items else ""
        block.append({
            "cells": [
                {"value": "Средний балл секции", "style": 14, "span": 4},
                {"value": fmt_score(avg), "style": score_style(avg), "span": 1},
                {"value": note or "Оценка секции по сигналам анализа речи.", "style": 5, "span": 1},
            ],
            "height": 24,
        })
        block.append({"cells": [], "height": 6})
        return block

    total_questions = len(questions_fragments)
    total_answers = len(answers_fragments)
    total_examples = len(examples_fragments)
    total_events = len(timeline_items)
    has_timeline = total_events > 0
    open_question_count = count_types(questions_fragments, {"open_ended", "clarifying", "factual"})
    checking_question_count = count_types(questions_fragments, {"checking_understanding", "quiz"})
    rhetorical_question_count = count_types(questions_fragments, {"rhetorical"})
    has_intro = bool(goals.get("intro", {}).get("present")) if isinstance(goals.get("intro"), dict) else False
    has_summary = bool(goals.get("summary", {}).get("present")) if isinstance(goals.get("summary"), dict) else False
    has_profanity = bool(profanity_block.get("present"))
    has_familiarity = bool(familiarity_block.get("present"))

    created_at = generation.get("created_at")
    date_text = ""
    if created_at:
        try:
            date_text = datetime.fromisoformat(str(created_at).replace("Z", "+00:00")).strftime("%d.%m.%Y")
        except Exception:
            date_text = str(created_at)

    def first_text(items: list[dict[str, Any]], fallback: str) -> str:
        for item in items:
            if not isinstance(item, dict):
                continue
            text = str(item.get("text") or "").strip()
            if text:
                return text
        return fallback

    rows: list[dict[str, Any]] = [
        {"cells": [{"value": "Чек-лист качества преподавания на занятии", "style": 2, "span": 6}], "height": 24},
        {"cells": [{"value": "Сделайте копию шаблона и заполните показатели на основе анализа речи преподавателя. Оставьте пустые ячейки там, где сигналов недостаточно для объективной оценки.", "style": 7, "span": 6}], "height": 36},
        {"cells": [{"value": "Дата формирования", "style": 14, "span": 1}, {"value": date_text, "style": 5, "span": 2}, {"value": "Файл", "style": 14, "span": 1}, {"value": str(generation.get("file_name") or "Без названия").strip(), "style": 5, "span": 2}], "height": 20},
        {"cells": [{"value": "Формат занятия", "style": 14, "span": 1}, {"value": str(lesson_format.get("format") or "Формат занятия не определен").strip(), "style": 5, "span": 2}, {"value": "Рекомендация", "style": 14, "span": 1}, {"value": str(teacher_recommendation.get("title") or "Рекомендация преподавателю").strip(), "style": 5, "span": 2}], "height": 20},
        {"cells": [{"value": "Комментарий", "style": 14, "span": 1}, {"value": str(lesson_format.get("comment") or "Агрегированный анализ речи преподавателя готов.").strip(), "style": 5, "span": 5}], "height": 28},
        {"cells": [], "height": 6},
        {"cells": [{"value": "Легенда", "style": 4, "span": 6}], "height": 20},
        {"cells": [{"value": "Баллы", "style": 1}, {"value": "Смысл", "style": 1}, {"value": "Комментарий", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}], "height": 20},
        {"cells": [{"value": "0", "style": 11, "span": 1}, {"value": "Отсутствие или несоответствие требованиям", "style": 5, "span": 2}, {"value": "Если показатель нельзя подтвердить по анализу или наблюдается нарушение.", "style": 7, "span": 3}], "height": 22},
        {"cells": [{"value": "1", "style": 12, "span": 1}, {"value": "Частичное соответствие", "style": 5, "span": 2}, {"value": "Сигнал есть, но он неполный, редкий или выражен неустойчиво.", "style": 7, "span": 3}], "height": 22},
        {"cells": [{"value": "2", "style": 13, "span": 1}, {"value": "Полное соответствие", "style": 5, "span": 2}, {"value": "Показатель проявляется уверенно и повторяется в анализе речи.", "style": 7, "span": 3}], "height": 22},
        {"cells": [], "height": 8},
        {"cells": [{"value": "Анализ по индикаторам", "style": 4, "span": 6}], "height": 20},
    ]

    rows.extend(section_rows("Организация понятного процесса обучения", 8, [
        {
            "competence": "Организация обучения",
            "indicator": "Приветствие и вход в урок",
            "manifestation": first_text(timeline_items, "В начале урока фиксируется вход в тему"),
            "score": 2 if has_intro or has_timeline else 1,
            "comment": "В начале есть структурирующий сигнал о входе в занятие." if has_intro or has_timeline else "Явного сигнала входа в урок не видно.",
        },
        {
            "competence": "Постановка задач",
            "indicator": "Формулирует цели и задачи занятия",
            "manifestation": str(goals.get("intro", {}).get("comment") or "Цели занятия проговариваются").strip() if isinstance(goals.get("intro"), dict) else "Цели занятия проговариваются",
            "score": 2 if has_intro else 0,
            "comment": "В начале урока есть явный сигнал о целях и рамке." if has_intro else "Явного сигнала о целях занятия не видно.",
        },
        {
            "competence": "Тайм-менеджмент",
            "indicator": "Укладывается в запланированную структуру",
            "manifestation": f"Этапов урока: {total_events}",
            "score": 2 if total_events >= 2 else 1 if total_events else 0,
            "comment": "В таймлайне просматриваются этапы занятия." if total_events else "Структура по времени выражена слабо.",
        },
        {
            "competence": "Педагогическая гибкость",
            "indicator": "Меняет подачу по ситуации",
            "manifestation": f"Примеры и аналогии: {total_examples}",
            "score": 2 if total_examples else 1 if total_questions else 0,
            "comment": "Примеры и аналогии помогают адаптировать объяснение." if total_examples else "Данных о гибкой перестройке подачи немного.",
        },
    ]))

    rows.extend(section_rows("Этика преподавания", 8, [
        {
            "competence": "Уважение к студентам и коллегам",
            "indicator": "Справедливо и объективно относится к участникам",
            "manifestation": "Ненормативная лексика и панибратство не подтверждены" if not has_profanity and not has_familiarity else "Есть сигналы, требующие внимания",
            "score": 2 if not has_profanity and not has_familiarity else 0,
            "comment": "Тон выглядит профессиональным и дистанция выдержана." if not has_profanity and not has_familiarity else "Есть речевые сигналы для доработки.",
        },
        {
            "competence": "Создание благоприятной атмосферы",
            "indicator": "Студенты не боятся задавать вопросы и комментировать",
            "manifestation": f"Вопросов преподавателя: {total_questions}, ответов студентов: {total_answers}",
            "score": 2 if total_questions and total_answers and not has_profanity else 1 if total_questions else 0,
            "comment": "Есть диалог и ответы студентов, атмосфера поддерживается вопросами." if total_questions else "Для уверенного вывода не хватает признаков диалога.",
        },
    ]))

    rows.extend(section_rows("Качество материала и владение им", 8, [
        {
            "competence": "Подготовка к занятию",
            "indicator": "Понимает образовательные результаты и структуру материала",
            "manifestation": str(lesson_format.get("format") or "Структурированный анализ речи").strip(),
            "score": 2 if lesson_format else 1,
            "comment": str(lesson_format.get("comment") or "Анализ построен по структуре урока.").strip(),
        },
        {
            "competence": "Доходчивость",
            "indicator": "Объясняет сложные моменты доступным способом",
            "manifestation": f"Примеры/аналогии: {total_examples}; ответы студентов: {total_answers}",
            "score": 2 if total_examples else 1 if total_answers else 0,
            "comment": "Объяснение поддерживается примерами и короткими пояснениями." if total_examples else "Пока мало прямых сигналов о доступности объяснения.",
        },
        {
            "competence": "Актуальность и широта",
            "indicator": "Выходит за рамки сухого пересказа",
            "manifestation": first_text(examples_fragments, "Примеры и аналогии как контекстуализация материала"),
            "score": 2 if total_examples >= 2 else 1 if total_examples else 0,
            "comment": "Материал подаётся через живые примеры и контекст." if total_examples else "Сигналов о широте контекста немного.",
        },
        {
            "competence": "Реакция на вопросы",
            "indicator": "Отвечает и развивает ответы аудитории",
            "manifestation": f"Вопросы: {total_questions}, ответы: {total_answers}",
            "score": 2 if total_answers else 1 if total_questions else 0,
            "comment": "В анализе есть взаимодействие на вопрос-ответ." if total_answers else "Недостаточно ответных реплик студентов.",
        },
    ]))

    rows.extend(section_rows("Управление динамикой занятия", 8, [
        {
            "competence": "Организация взаимодействия",
            "indicator": "Вовлекает студентов в процесс обучения",
            "manifestation": f"Всего вопросов: {total_questions}",
            "score": 2 if total_questions else 0,
            "comment": f"Есть {total_questions} вопросов к аудитории." if total_questions else "Вопросов к аудитории не найдено.",
        },
        {
            "competence": "Умение слушать",
            "indicator": "Серьёзно подходит к ответам студентов",
            "manifestation": f"Ответов студентов: {total_answers}",
            "score": 2 if total_answers else 0,
            "comment": "Есть ответы студентов, на которые можно опереться." if total_answers else "В данных нет ответов студентов.",
        },
        {
            "competence": "Мониторинг понимания",
            "indicator": "Регулярно проверяет понимание материала",
            "manifestation": f"Вопросов на проверку понимания: {checking_question_count}",
            "score": 2 if checking_question_count else 1 if total_questions else 0,
            "comment": "Вопросы на понимание присутствуют." if checking_question_count else "Проверка понимания выражена слабо.",
        },
        {
            "competence": "Обратная связь",
            "indicator": "Даёт конструктивную обратную связь",
            "manifestation": "Ответы и уточнения фиксируются в анализе" if total_answers else "Сигналов обратной связи мало",
            "score": 2 if total_answers else 0,
            "comment": "Взаимодействие с ответами студентов есть." if total_answers else "Недостаточно данных для оценки обратной связи.",
        },
    ]))

    rows.extend(section_rows("Структура речи и языка", 8, [
        {
            "competence": "Структурирование презентации",
            "indicator": "Чётко и логично выстраивает материал",
            "manifestation": f"Этапов урока: {total_events}",
            "score": 2 if total_events else 1,
            "comment": "Таймлайн показывает последовательность изложения." if total_events else "Структура изложения просматривается слабо.",
        },
        {
            "competence": "Грамотность речи",
            "indicator": "Избегает слов-паразитов и грубых выражений",
            "manifestation": "Ненормативная лексика не подтверждена" if not has_profanity else "Есть сигналы на проверку",
            "score": 2 if not has_profanity else 0,
            "comment": "Речь выглядит аккуратной и профессиональной." if not has_profanity else "Есть сигналы на доработку речевой культуры.",
        },
        {
            "competence": "Темп речи",
            "indicator": "Держит темп, при котором студенты успевают воспринимать информацию",
            "manifestation": f"Событий урока: {total_events}; вопросных сигналов: {total_questions}",
            "score": 2 if total_events >= 2 else 1,
            "comment": "Есть распределение по этапам и паузы для проверки понимания." if total_events >= 2 else "Темп трудно оценить по имеющимся сигналам.",
        },
        {
            "competence": "Понятные языковые конструкции",
            "indicator": "Мысли выражаются ёмко и ясно",
            "manifestation": str(lesson_format.get("comment") or "Наблюдается структурированная подача").strip(),
            "score": 2 if total_examples or total_questions else 1,
            "comment": "Формулировки выглядят собранными и понятными." if total_examples or total_questions else "Сигналов о языковой ясности немного.",
        },
    ]))

    rows.extend(section_rows("Приёмы вовлечения аудитории", 8, [
        {
            "competence": "Управление вниманием аудитории",
            "indicator": "Задаёт вопросы, чтобы вовлечь студентов",
            "manifestation": f"Всего вопросов: {total_questions}",
            "score": 2 if total_questions else 0,
            "comment": "Вовлечение строится через вопросы к аудитории." if total_questions else "Вопросные механики не просматриваются.",
        },
        {
            "competence": "Открытые вопросы",
            "indicator": "Использует открытые/уточняющие вопросы",
            "manifestation": f"Открытых и уточняющих вопросов: {open_question_count}",
            "score": 2 if open_question_count else 1 if total_questions else 0,
            "comment": "Вопросы на разворачивание мысли присутствуют." if open_question_count else "Открытые вопросы не доминируют.",
        },
        {
            "competence": "Примеры и аналогии",
            "indicator": "Использует сторителлинг, примеры и аналогии",
            "manifestation": f"Примеров/аналогий: {total_examples}",
            "score": 2 if total_examples else 0,
            "comment": "Примеры и аналогии помогают удерживать внимание." if total_examples else "Примеры и аналогии не отмечены.",
        },
        {
            "competence": "Разнообразие механик",
            "indicator": "Использует несколько способов вовлечения",
            "manifestation": f"Вопросы: {total_questions}, примеры: {total_examples}, риторические вопросы: {rhetorical_question_count}",
            "score": 2 if total_questions and total_examples else 1 if total_questions or total_examples else 0,
            "comment": "В анализе виден микс вопросов и примеров." if total_questions and total_examples else "Набор приёмов пока ограничен.",
        },
    ]))

    rows.extend(section_rows("Итог и рекомендация", 8, [
        {
            "competence": "Рекомендация преподавателю",
            "indicator": "Итоговый комментарий модели",
            "manifestation": str(teacher_recommendation.get("title") or "Рекомендация преподавателю").strip(),
            "score": 2 if str(teacher_recommendation.get("comment") or "").strip() else 1,
            "comment": str(teacher_recommendation.get("comment") or "Комментарий отсутствует.").strip(),
            "summary_note": "Это финальный вывод по всем сигналам анализа речи.",
        },
    ]))

    numeric_scores: list[float] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        cells = row.get("cells") if isinstance(row.get("cells"), list) else []
        for cell in cells:
            if isinstance(cell, dict) and int(cell.get("style", 0) or 0) in {11, 12, 13}:
                try:
                    numeric_scores.append(float(str(cell.get("value") or 0).replace(",", ".")))
                except Exception:
                    pass

    overall_score = sum(numeric_scores) / len(numeric_scores) if numeric_scores else 0
    rows.append({"cells": [{"value": "Итоговый средний балл", "style": 14, "span": 4}, {"value": "-", "style": score_style(overall_score), "span": 1}, {"value": "Сводная оценка по всем индикаторам анализа речи.", "style": 5, "span": 1}], "height": 28})
    rows.append({"cells": [], "height": 6})

    return [
        {
            "name": "Анализ речи",
            "rows": rows,
            "cols": [20, 28, 30, 16, 12, 40],
            "page_setup": {"orientation": "landscape", "paperSize": 9, "fitToWidth": 1, "fitToHeight": 1},
            "freeze_panes": "A12",
        }
    ]


def build_speech_analysis_export_worksheets_from_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Builds worksheets formatted for teacher report download starting from raw speech analysis dict payload."""
    transcript = payload.get("transcript") if isinstance(payload.get("transcript"), list) else []
    speech = payload.get("speech_analysis") if isinstance(payload.get("speech_analysis"), dict) else {}
    if not speech:
        return []
    generation = {
        "transcript": transcript,
        "analytics": {"speech_analysis": speech},
    }
    return build_speech_analysis_export_worksheets_precise(generation)


def build_xlsx_bytes(worksheets: list[dict[str, Any]]) -> bytes:
    """Takes structured sheets and converts them to binary .xlsx bytes via openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    unique_sheets: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for index, sheet in enumerate(worksheets, start=1):
        base_name = sanitize_xlsx_sheet_name(str(sheet.get("name") or f"Лист {index}"))
        name = base_name
        suffix = 2
        while name in used_names:
            trimmed = base_name[: max(1, 31 - len(f" ({suffix})"))].rstrip()
            name = f"{trimmed} ({suffix})"[:31]
            suffix += 1
        used_names.add(name)
        unique_sheets.append({
            "name": name,
            "rows": sheet.get("rows") if isinstance(sheet.get("rows"), list) else [],
            "cols": sheet.get("cols") if isinstance(sheet.get("cols"), list) else [],
            "page_setup": sheet.get("page_setup") if isinstance(sheet.get("page_setup"), dict) else {},
            "freeze_panes": sheet.get("freeze_panes") if isinstance(sheet.get("freeze_panes"), str) else "",
        })

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "FastClass"
    wb.properties.lastModifiedBy = "FastClass"

    thin = Side(style="thin", color="FFD9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")
    title_fill = PatternFill(fill_type="solid", fgColor="FFC7DDF4")
    note_fill = PatternFill(fill_type="solid", fgColor="FFF8EFB6")
    sub_fill = PatternFill(fill_type="solid", fgColor="FFF1F5F9")
    section_blue_fill = PatternFill(fill_type="solid", fgColor="FFE8F1FF")
    section_green_fill = PatternFill(fill_type="solid", fgColor="FFE9F9EE")
    section_purple_fill = PatternFill(fill_type="solid", fgColor="FFF0E9FF")
    section_orange_fill = PatternFill(fill_type="solid", fgColor="FFFFEFE3")
    score_red_fill = PatternFill(fill_type="solid", fgColor="FFE76F61")
    score_orange_fill = PatternFill(fill_type="solid", fgColor="FFF2B28F")
    score_green_fill = PatternFill(fill_type="solid", fgColor="FF87C88D")
    score_gray_fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    header_text_fill = PatternFill(fill_type="solid", fgColor="FFFCE7D2")

    styles = {
        0: {
            "font": Font(name="Arial", size=10, color="FF1E293B"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        1: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": header_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        2: {
            "font": Font(name="Arial", size=14, bold=True, color="FF0F172A"),
            "fill": title_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        3: {
            "font": Font(name="Arial", size=10, italic=True, color="FF64748B"),
            "fill": note_fill,
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        4: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": sub_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        5: {
            "font": Font(name="Arial", size=10, color="FF1E293B"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        6: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": header_text_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        7: {
            "font": Font(name="Arial", size=9, italic=True, color="FF1F2937"),
            "fill": note_fill,
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        8: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_blue_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        9: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_green_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        10: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_purple_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        11: {
            "font": Font(name="Arial", size=11, bold=True, color="FFFFFFFF"),
            "fill": score_red_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        12: {
            "font": Font(name="Arial", size=11, bold=True, color="FF1F2937"),
            "fill": score_orange_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        13: {
            "font": Font(name="Arial", size=11, bold=True, color="FFFFFFFF"),
            "fill": score_green_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        14: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": score_gray_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        15: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
    }

    def apply_style(cell, style_id: Optional[int]) -> None:
        spec = styles.get(int(style_id) if style_id is not None else 0, styles[0])
        cell.font = spec["font"]
        cell.fill = spec["fill"]
        cell.alignment = spec["alignment"]
        cell.border = border

    def sanitize_xlsx_value(value: Any) -> Any:
        if isinstance(value, str):
            return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
        if isinstance(value, (int, float, bool)) or value is None:
            return value
        return sanitize_xlsx_value(str(value))

    for sheet in unique_sheets:
        ws = wb.create_sheet(title=sheet["name"])
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        rows = sheet["rows"]
        max_cols = max(6, len(sheet.get("cols") or []))
        text_widths = [12.0] * max_cols

        for row_num, row in enumerate(rows, start=1):
            row_spec = row if isinstance(row, dict) else {"cells": row if isinstance(row, list) else [row], "height": None}
            cells = row_spec.get("cells")
            if not isinstance(cells, list):
                cells = [cells] if cells is not None else []
            if row_spec.get("height") is not None:
                try:
                    ws.row_dimensions[row_num].height = float(row_spec["height"])
                except (TypeError, ValueError):
                    pass

            col_num = 1
            for cell in cells:
                if isinstance(cell, dict):
                    value = cell.get("value", "")
                    span_raw = cell.get("span", 1)
                    try:
                        span = max(1, int(span_raw or 1))
                    except (TypeError, ValueError):
                        span = 1
                    style_id = cell.get("style")
                else:
                    value = cell
                    span = 1
                    style_id = None

                value = sanitize_xlsx_value(value)

                if col_num > max_cols:
                    max_cols = col_num
                    text_widths.extend([12.0] * (max_cols - len(text_widths)))

                top_left = ws.cell(row=row_num, column=col_num, value=value)
                apply_style(top_left, style_id)

                if span > 1:
                    end_col = col_num + span - 1
                    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=end_col)
                    max_cols = max(max_cols, end_col)
                    if len(text_widths) < max_cols:
                        text_widths.extend([12.0] * (max_cols - len(text_widths)))
                value_text = str(value or "").replace("\n", " ")
                estimated = min(42.0, max(12.0, (len(value_text) * 0.85) + 2.0))
                if span == 1:
                    idx = col_num - 1
                    text_widths[idx] = max(text_widths[idx], estimated)
                else:
                    span_width = estimated / span
                    for offset in range(span):
                        idx = col_num - 1 + offset
                        if idx >= len(text_widths):
                            text_widths.extend([12.0] * (idx - len(text_widths) + 1))
                        text_widths[idx] = max(text_widths[idx], min(28.0, span_width))
                col_num += span

        for col_idx in range(1, max_cols + 1):
            width = text_widths[col_idx - 1] if col_idx - 1 < len(text_widths) else 12.0
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10.0), 42.0)

        freeze_panes = sheet.get("freeze_panes") or "A5"
        ws.freeze_panes = freeze_panes

    return _write_workbook_to_bytes(wb)


def _write_workbook_to_bytes(workbook: Any) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
